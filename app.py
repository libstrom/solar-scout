"""Enspecta Solar Lead Machine -- Streamlit UI.

Tabs:
  Dashboard         -- counts, status chart, live scan progress + ETA.
  Verification Lab  -- one rooftop at a time, best AI score first,
                       GREEN (Confirm) / RED (Reject).
  Ringlista         -- confirmed leads: phone/owner/call status, saved to DB.
  Export            -- download confirmed leads as Excel.

Run:
    streamlit run app.py
"""
from __future__ import annotations

import io
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote_plus

import pandas as pd
import streamlit as st
from openlocationcode import openlocationcode as olc
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import db as shared_db
from harvester import format_eta

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "data" / "leads.db"

CALL_STATUSES = ["", "Att ringa", "Uppringd", "Bokad", "Nej tack"]


# ---- DB --------------------------------------------------------------------

def _conn() -> sqlite3.Connection:
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    return c


def stats() -> dict:
    if not DB_PATH.exists():
        return {"total": 0, "pending": 0, "confirmed": 0, "rejected": 0}
    with _conn() as c:
        out = {"total": c.execute("SELECT COUNT(*) FROM leads").fetchone()[0]}
        for s in ("pending", "confirmed", "rejected"):
            out[s] = c.execute(
                "SELECT COUNT(*) FROM leads WHERE status = ?", (s,)
            ).fetchone()[0]
    return out


def list_pending() -> list[sqlite3.Row]:
    if not DB_PATH.exists():
        return []
    with _conn() as c:
        # Best AI score first; ungraded last in harvest order.
        return c.execute(
            """
            SELECT * FROM leads WHERE status = 'pending'
            ORDER BY (ai_score IS NULL), ai_score DESC, id ASC
            """
        ).fetchall()


def list_confirmed() -> list[sqlite3.Row]:
    if not DB_PATH.exists():
        return []
    with _conn() as c:
        return c.execute(
            "SELECT * FROM leads WHERE status = 'confirmed' ORDER BY id ASC"
        ).fetchall()


def set_status(lead_id: int, status: str) -> None:
    if status not in ("confirmed", "rejected"):
        raise ValueError(status)
    with _conn() as c:
        c.execute(
            "UPDATE leads SET status = ?, verified_at = ? WHERE id = ?",
            (status,
             datetime.now(timezone.utc).replace(tzinfo=None).isoformat(timespec="seconds"),
             lead_id),
        )
        c.commit()


def save_call_fields(lead_id: int, phone: str, owner: str, call_status: str, notes: str) -> None:
    with _conn() as c:
        c.execute(
            "UPDATE leads SET phone = ?, owner_name = ?, call_status = ?, call_notes = ? WHERE id = ?",
            (phone or None, owner or None, call_status or None, notes or None, lead_id),
        )
        c.commit()


# ---- Computed export-time fields ------------------------------------------

def estimate_fuse(area_m2: float | None) -> str:
    """Heuristic Swedish residential service fuse based on roof area."""
    if area_m2 is None:
        return ""
    if area_m2 < 100:
        return "16 A"
    if area_m2 < 150:
        return "20 A"
    if area_m2 < 200:
        return "25 A"
    return "35 A"


def plus_code(lat: float, lng: float) -> str:
    return olc.encode(lat, lng)


def mrkoll_link(address: str | None) -> str:
    if not address:
        return ""
    return f"https://mrkoll.se/resultat/?n=&a={quote_plus(address)}"


def maps_link(lat: float, lng: float) -> str:
    return f"https://www.google.com/maps/search/?api=1&query={lat},{lng}"


# ---- Excel builder ---------------------------------------------------------

EXPORT_COLUMNS = [
    "Address",
    "Plus Code",
    "Estimated Fuse (A)",
    "Telefon",
    "Ägare",
    "Ringstatus",
    "AI Score",
    "MrKoll Link",
    "Maps Link",
]


def build_excel_bytes(rows: list[sqlite3.Row]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Confirmed Leads"
    ws.append(EXPORT_COLUMNS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    keys = rows[0].keys() if rows else []
    for r in rows:
        ws.append([
            r["address"] or "",
            plus_code(r["lat"], r["lng"]),
            estimate_fuse(r["roof_area_m2"]),
            (r["phone"] if "phone" in keys else "") or "",
            (r["owner_name"] if "owner_name" in keys else "") or "",
            (r["call_status"] if "call_status" in keys else "") or "",
            (r["ai_score"] if "ai_score" in keys else "") or "",
            mrkoll_link(r["address"]),
            maps_link(r["lat"], r["lng"]),
        ])

    widths = {"A": 42, "B": 18, "C": 16, "D": 16, "E": 22, "F": 12, "G": 10, "H": 60, "I": 60}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---- UI --------------------------------------------------------------------

st.set_page_config(page_title="Enspecta Solar Lead Machine", layout="wide")
st.title("Enspecta Solar Lead Machine")

if not DB_PATH.exists():
    st.warning(
        "No database found yet. Run the harvester first:\n\n"
        "```\npython harvester.py\n```"
    )
    st.stop()

shared_db.ensure_schema()

tab_dashboard, tab_verify, tab_calls, tab_export = st.tabs([
    "Dashboard", "Verification Lab", "Ringlista", "Export",
])


# ---- Dashboard tab ---------------------------------------------------------

def _render_scan_panel() -> None:
    run = shared_db.latest_scan_run()
    if run is None:
        st.caption("Ingen scan körd ännu. Starta med `python harvester.py`.")
        return

    pct = (run["grid_done"] / run["grid_total"]) if run["grid_total"] else 0.0
    if run["status"] == "running":
        st.markdown("#### Pågående scan: " + run["town"])
        st.progress(min(1.0, pct),
                    text=f"{run['grid_done']:,}/{run['grid_total']:,} grid-punkter ({pct:.1%})")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Nya leads", f"{run['new_leads']}/{run['max_buildings']}")
        c2.metric("ETA", format_eta(run["eta_seconds"]))
        c3.metric("Kostnad hittills", f"${run['cost_usd']:.2f}")
        c4.metric("Fel", run["errors"])
        st.caption(f"Uppdaterad {run['updated_at'] or run['started_at']} UTC")
    else:
        label = {"done": "klar", "aborted": "avbruten", "error": "fel"}.get(run["status"], run["status"])
        st.caption(
            f"Senaste scan ({run['town']}, {label}): "
            f"{run['new_leads']} nya leads, {run['grid_done']:,}/{run['grid_total']:,} punkter, "
            f"~${run['cost_usd']:.2f}, startad {run['started_at']} UTC."
        )


if hasattr(st, "fragment"):
    _render_scan_panel = st.fragment(run_every="3s")(_render_scan_panel)

with tab_dashboard:
    _render_scan_panel()
    st.divider()

    s = stats()
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total scanned", s["total"])
    c2.metric("Pending", s["pending"])
    c3.metric("Confirmed", s["confirmed"])
    c4.metric("Rejected", s["rejected"])

    if s["total"]:
        df = pd.DataFrame({
            "Status": ["Pending", "Confirmed", "Rejected"],
            "Count":  [s["pending"], s["confirmed"], s["rejected"]],
        })
        st.bar_chart(df.set_index("Status"))
    else:
        st.info("No leads in the DB yet. Run `python harvester.py` to populate.")


# ---- Verification Lab tab --------------------------------------------------

with tab_verify:
    pending = list_pending()
    if not pending:
        st.success("No pending leads. Either all done -- or nothing harvested yet.")
    else:
        graded = sum(1 for r in pending if r["ai_score"] is not None)
        if graded < len(pending):
            st.caption(
                f"AI-graderade: {graded}/{len(pending)} pending. "
                "Kör `python prescreen.py` för att sortera bästa taken först."
            )

        idx = st.session_state.setdefault("verify_idx", 0)
        idx = max(0, min(idx, len(pending) - 1))
        row = pending[idx]

        left, right = st.columns([2, 1])
        with left:
            st.subheader(f"Lead {idx + 1} of {len(pending)} pending")
            if row["ai_score"] is not None:
                if row["ai_has_panels"]:
                    st.error(f"AI: SOLCELLER FINNS REDAN — {row['ai_reason'] or ''}")
                elif row["ai_score"] >= 70:
                    st.success(f"AI score {row['ai_score']}/100 — {row['ai_reason'] or ''}")
                else:
                    st.info(f"AI score {row['ai_score']}/100 — {row['ai_reason'] or ''}")
            if row["image_path"] and Path(row["image_path"]).exists():
                st.image(row["image_path"],
                         caption=row["address"] or "(no address)",
                         use_container_width=True)
            else:
                st.warning(f"Image missing: {row['image_path']}")

        with right:
            st.markdown("### Metadata")
            st.write(f"**Address:** {row['address'] or '—'}")
            st.write(f"**Coordinates:** `{row['coordinates']}`")
            st.write(f"**Roof area:** {row['roof_area_m2'] or '?'} m²")
            st.write(f"**Solar confidence:** `{row['solar_confidence']}`")
            st.write(f"**Plus Code:** `{plus_code(row['lat'], row['lng'])}`")
            st.write(f"**Estimated fuse (heuristic):** {estimate_fuse(row['roof_area_m2']) or '—'}")
            st.write(f"**place_id:** `{row['place_id']}`")

            st.divider()
            g, r = st.columns(2)
            with g:
                if st.button("CONFIRM (Green)", type="primary", use_container_width=True):
                    set_status(row["id"], "confirmed")
                    st.session_state["verify_idx"] = idx
                    st.rerun()
            with r:
                if st.button("REJECT (Red)", use_container_width=True):
                    set_status(row["id"], "rejected")
                    st.session_state["verify_idx"] = idx
                    st.rerun()

            n1, n2, n3 = st.columns(3)
            with n1:
                if st.button("Prev"):
                    st.session_state["verify_idx"] = max(0, idx - 1)
                    st.rerun()
            with n2:
                if st.button("Skip"):
                    st.session_state["verify_idx"] = min(len(pending) - 1, idx + 1)
                    st.rerun()
            with n3:
                if st.button("Next"):
                    st.session_state["verify_idx"] = min(len(pending) - 1, idx + 1)
                    st.rerun()


# ---- Ringlista tab -----------------------------------------------------------

with tab_calls:
    confirmed = list_confirmed()
    if not confirmed:
        st.info("Inga bekräftade leads ännu. Bekräfta tak i Verification Lab först.")
    else:
        st.write(
            f"**{len(confirmed)}** bekräftade leads. Slå upp ägare/telefon via "
            "MrKoll-länken, fyll i och spara — statusen följer med i exporten."
        )
        keys = confirmed[0].keys()
        df = pd.DataFrame([{
            "id": r["id"],
            "Adress": r["address"] or "",
            "AI": (r["ai_score"] if "ai_score" in keys else None),
            "Telefon": (r["phone"] if "phone" in keys else "") or "",
            "Ägare": (r["owner_name"] if "owner_name" in keys else "") or "",
            "Status": (r["call_status"] if "call_status" in keys else "") or "",
            "Anteckningar": (r["call_notes"] if "call_notes" in keys else "") or "",
            "MrKoll": mrkoll_link(r["address"]),
        } for r in confirmed])

        edited = st.data_editor(
            df,
            hide_index=True,
            use_container_width=True,
            disabled=["id", "Adress", "AI", "MrKoll"],
            column_config={
                "id": st.column_config.NumberColumn(width="small"),
                "AI": st.column_config.NumberColumn(width="small"),
                "Status": st.column_config.SelectboxColumn(options=CALL_STATUSES),
                "MrKoll": st.column_config.LinkColumn(display_text="MrKoll"),
            },
            key="ringlista_editor",
        )

        if st.button("Spara ringlistan", type="primary"):
            for _, rrow in edited.iterrows():
                save_call_fields(
                    int(rrow["id"]),
                    str(rrow["Telefon"]).strip(),
                    str(rrow["Ägare"]).strip(),
                    str(rrow["Status"]).strip(),
                    str(rrow["Anteckningar"]).strip(),
                )
            st.success("Sparat.")
            st.rerun()

        booked = sum(1 for _, x in edited.iterrows() if x["Status"] == "Bokad")
        to_call = sum(1 for _, x in edited.iterrows() if x["Status"] == "Att ringa")
        c1, c2 = st.columns(2)
        c1.metric("Att ringa", to_call)
        c2.metric("Bokade möten", booked)


# ---- Export tab ------------------------------------------------------------

with tab_export:
    confirmed = list_confirmed()
    st.write(f"**{len(confirmed)}** confirmed leads ready for export.")
    if confirmed:
        keys = confirmed[0].keys()
        preview = pd.DataFrame([{
            "Address": r["address"] or "",
            "Plus Code": plus_code(r["lat"], r["lng"]),
            "Estimated Fuse (A)": estimate_fuse(r["roof_area_m2"]),
            "Telefon": (r["phone"] if "phone" in keys else "") or "",
            "Ringstatus": (r["call_status"] if "call_status" in keys else "") or "",
            "MrKoll Link": mrkoll_link(r["address"]),
            "Maps Link": maps_link(r["lat"], r["lng"]),
        } for r in confirmed])
        st.dataframe(preview, use_container_width=True)

        xlsx_bytes = build_excel_bytes(confirmed)
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        st.download_button(
            label="Download Enspecta_Leads.xlsx",
            data=xlsx_bytes,
            file_name=f"Enspecta_Leads_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )
    else:
        st.info("Confirm some leads in the Verification Lab to enable export.")
