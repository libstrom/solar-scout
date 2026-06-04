#!/usr/bin/env python3
"""
exportEnergyList.py — Exporterar alla fastigheter ur energy-data.json till Excel/CSV
                       för manuell sökning i FileMaker eller annat CRM.

Usage:
    python exportEnergyList.py energy-data.json [output.xlsx]
    python exportEnergyList.py energy-data.json output.csv

Output-fält (per rad):
  Fastighetsbeteckning, Energiklass, kWh/m², Byggår, Atemp m², Har sol,
  Har solvärme, Uppvärmning, Adress, Postnr, Ort, Kommun,
  Deklarationsdatum, El direkt kWh, Källa (xlsm/pdf), Source-fil
"""
import sys
import json
import re
from pathlib import Path
from datetime import date

ENERGIKLASS_ORDER = {'A': 0, 'B': 1, 'C': 2, 'D': 3, 'E': 4, 'F': 5, 'G': 6, '': 9}


def ek_sort(rec):
    ek = (rec.get('energiklass') or '').upper().strip()
    return ENERGIKLASS_ORDER.get(ek, 8)


def epk_el(epk):
    if not epk or not isinstance(epk, dict):
        return 0
    return (epk.get('el_direkt') or 0) + (epk.get('el_vattenburen') or 0)


def tc(s):
    if not s:
        return ''
    return str(s).strip().title()


def build_rows(data: dict) -> list[dict]:
    rows = []
    for fastig, rec in data.items():
        ek = (rec.get('energiklass') or '').upper().strip()
        epk = rec.get('energi_per_kalla') or {}
        rows.append({
            'Fastighetsbeteckning': tc(fastig),
            'Energiklass':          ek,
            'Energiprestanda kWh/m²': rec.get('energiprestanda_kwh') or '',
            'Byggår':               rec.get('nybyggnadsar') or '',
            'Atemp m²':             rec.get('atemp_m2') or '',
            'Har solceller':        'Ja' if rec.get('har_solceller') else '',
            'Har solvärme':         'Ja' if rec.get('har_solvarme') else '',
            'Uppvärmning':          rec.get('uppvarmningssystem') or '',
            'Adress':               rec.get('adress') or '',
            'Postnr':               rec.get('postnummer') or rec.get('postnr') or '',
            'Ort':                  tc(rec.get('ort') or ''),
            'Kommun':               tc(rec.get('kommun') or ''),
            'Deklarationsdatum':    rec.get('deklaration_datum') or '',
            'El direkt kWh':        epk_el(epk) or '',
            'Källa':                rec.get('source') or '',
            'Source-fil':           rec.get('source_file') or '',
        })
    # Sort: energiklass A→G, then fastig
    rows.sort(key=lambda r: (ENERGIKLASS_ORDER.get(r['Energiklass'], 8), r['Fastighetsbeteckning']))
    return rows


def write_excel(rows: list[dict], out_path: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        print('Kör: pip install openpyxl')
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Energideklarationer'
    ws.sheet_view.showGridLines = False

    headers = list(rows[0].keys()) if rows else []
    COL_WIDTHS = {
        'Fastighetsbeteckning': 26,
        'Energiklass': 10,
        'Energiprestanda kWh/m²': 18,
        'Byggår': 8,
        'Atemp m²': 9,
        'Har solceller': 12,
        'Har solvärme': 11,
        'Uppvärmning': 22,
        'Adress': 28,
        'Postnr': 8,
        'Ort': 16,
        'Kommun': 16,
        'Deklarationsdatum': 16,
        'El direkt kWh': 13,
        'Källa': 8,
        'Source-fil': 50,
    }

    EK_COLORS = {
        'A': ('1B5E20', 'E8F5E9'),
        'B': ('2E7D32', 'F1F8E9'),
        'C': ('558B2F', 'F9FBE7'),
        'D': ('F9A825', 'FFFDE7'),
        'E': ('E65100', 'FFF3E0'),
        'F': ('C62828', 'FFEBEE'),
        'G': ('B71C1C', 'FCE4EC'),
    }

    hdr_fill = PatternFill('solid', fgColor='1A237E')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    thin = Side(style='thin', color='DDDDDD')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(h, 14)
    ws.row_dimensions[1].height = 22

    # Data rows
    for ri, row in enumerate(rows, 2):
        ek = row.get('Energiklass', '')
        fg, bg = EK_COLORS.get(ek, ('333333', 'FFFFFF'))
        row_fill = PatternFill('solid', fgColor=bg)
        for ci, h in enumerate(headers, 1):
            val = row[h]
            c = ws.cell(ri, ci, val)
            c.border = bdr
            c.font = Font(size=9)
            c.alignment = Alignment(vertical='center')
            if h == 'Energiklass' and ek:
                c.fill = PatternFill('solid', fgColor=EK_COLORS.get(ek, ('333333','FFFFFF'))[1])
                c.font = Font(bold=True, size=10, color=fg)
                c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                c.fill = row_fill if ek else PatternFill('solid', fgColor='F5F5F5')
        ws.row_dimensions[ri].height = 16

    # Table
    if rows:
        last_col = get_column_letter(len(headers))
        last_row = len(rows) + 1
        tbl = Table(displayName='EnergiLista', ref=f'A1:{last_col}{last_row}')
        tbl.tableStyleInfo = TableStyleInfo(
            name='TableStyleLight1', showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(tbl)

    ws.freeze_panes = 'A2'
    wb.save(out_path)
    print(f'Excel sparad: {out_path}  ({len(rows)} rader)')


def write_csv(rows: list[dict], out_path: str):
    import csv
    headers = list(rows[0].keys()) if rows else []
    with open(out_path, 'w', newline='', encoding='utf-8-sig') as f:  # utf-8-sig = BOM for Excel
        w = csv.DictWriter(f, fieldnames=headers, delimiter=';')
        w.writeheader()
        w.writerows(rows)
    print(f'CSV sparad: {out_path}  ({len(rows)} rader, semikolon-separerad, UTF-8 BOM)')


def main():
    args = sys.argv[1:]
    if not args:
        print('Usage: python exportEnergyList.py energy-data.json [output.xlsx|output.csv]')
        sys.exit(1)

    in_path  = args[0]
    out_path = args[1] if len(args) > 1 else 'energy-list.xlsx'

    print(f'Läser {in_path} ...')
    data = json.loads(Path(in_path).read_text(encoding='utf-8'))
    print(f'  {len(data)} fastigheter')

    rows = build_rows(data)

    has_sol = sum(1 for r in rows if r['Har solceller'])
    no_adr  = sum(1 for r in rows if not r['Adress'])
    ek_dist = {}
    for r in rows:
        ek_dist[r['Energiklass'] or '—'] = ek_dist.get(r['Energiklass'] or '—', 0) + 1

    print(f'  Har solceller: {has_sol}')
    print(f'  Utan adress:   {no_adr}  (beteckning finns alltid)')
    print(f'  Energiklass-fördelning: {dict(sorted(ek_dist.items()))}')
    print()

    ext = Path(out_path).suffix.lower()
    if ext == '.csv':
        write_csv(rows, out_path)
    else:
        write_excel(rows, out_path)


if __name__ == '__main__':
    main()
