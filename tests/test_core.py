"""Tests for the pure functions in harvester.py and app.py.

No network calls are made and the real data/leads.db is never opened:

* harvester DB tests redirect DB_PATH / DATA_DIR / IMAGES_DIR to tmp_path
  via monkeypatch before any DB function is called (db() reads the module
  globals at call time, so setattr on the module works).
* app.py runs Streamlit code at import time, so streamlit is replaced with
  a MagicMock in sys.modules BEFORE the import.  During the import,
  pathlib.Path.exists is patched to return False so app.py takes its
  "no database yet" branch: st.warning(); st.stop().  st.stop is given a
  side_effect exception that aborts module execution right after all the
  pure functions are defined -- the real DB is never queried.
"""
from __future__ import annotations

import importlib.util
import math
import sqlite3
import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

PROJECT_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_DIR))

import harvester  # noqa: E402  (no top-level side effects beyond load_dotenv)


# ---- app.py import with mocked streamlit ------------------------------------

class _StreamlitStop(Exception):
    """Raised by the mocked st.stop() to abort app.py's top-level UI code."""


def _import_app():
    st_mock = MagicMock(name="streamlit")
    st_mock.stop.side_effect = _StreamlitStop
    # Defensive config in case more top-level code runs than expected:
    st_mock.tabs.return_value = (MagicMock(), MagicMock(), MagicMock())
    st_mock.columns.side_effect = lambda spec: tuple(
        MagicMock() for _ in range(spec if isinstance(spec, int) else len(spec))
    )
    st_mock.button.return_value = False
    st_mock.session_state = {}
    sys.modules["streamlit"] = st_mock

    # Pre-import app.py's other dependencies so they are cached in
    # sys.modules and unaffected by the Path.exists patch below.
    import pandas  # noqa: F401
    import openpyxl  # noqa: F401
    from openlocationcode import openlocationcode  # noqa: F401

    spec = importlib.util.spec_from_file_location("app", PROJECT_DIR / "app.py")
    mod = importlib.util.module_from_spec(spec)
    sys.modules["app"] = mod
    # Path.exists -> False forces the st.warning()/st.stop() branch so the
    # real data/leads.db is never touched during import.
    with patch.object(Path, "exists", lambda self: False):
        with pytest.raises(_StreamlitStop):
            spec.loader.exec_module(mod)
    return mod


app = _import_app()

KAGEROD_BBOX = harvester.BBOXES["Kågeröd"]


# ---- harvester: grid maths ---------------------------------------------------

class TestMetresToDegrees:
    def test_latitude_degree_at_lat_56(self):
        d_lat, _ = harvester.metres_to_degrees(20, 56.0)
        assert d_lat == pytest.approx(20 / 111_320.0)

    def test_longitude_degree_stretched_by_cos_lat(self):
        d_lat, d_lng = harvester.metres_to_degrees(20, 56.0)
        assert d_lng == pytest.approx(d_lat / math.cos(math.radians(56.0)))
        # At lat 56, a longitude degree is ~62 km, so 20 m is ~0.00032 deg.
        assert 0.0003 < d_lng < 0.00034

    def test_zero_metres(self):
        assert harvester.metres_to_degrees(0, 56.0) == (0.0, 0.0)

    def test_scales_linearly(self):
        d1 = harvester.metres_to_degrees(10, 56.0)
        d2 = harvester.metres_to_degrees(20, 56.0)
        assert d2[0] == pytest.approx(2 * d1[0])
        assert d2[1] == pytest.approx(2 * d1[1])


class TestGridPoints:
    def test_count_matches_grid_size_for_kagerod(self):
        pts = list(harvester.grid_points(KAGEROD_BBOX, harvester.GRID_SPACING_M))
        assert len(pts) == harvester.grid_size(KAGEROD_BBOX, harvester.GRID_SPACING_M)

    def test_all_points_within_bbox(self):
        south, west, north, east = KAGEROD_BBOX
        for lat, lng in harvester.grid_points(KAGEROD_BBOX, harvester.GRID_SPACING_M):
            assert south <= lat <= north
            assert west <= lng <= east

    def test_first_point_is_southwest_corner(self):
        first = next(harvester.grid_points(KAGEROD_BBOX, harvester.GRID_SPACING_M))
        assert first == (KAGEROD_BBOX[0], KAGEROD_BBOX[1])


class TestGridSize:
    @pytest.mark.parametrize("town", list(harvester.BBOXES))
    def test_positive_for_all_configured_bboxes(self, town):
        assert harvester.grid_size(harvester.BBOXES[town], harvester.GRID_SPACING_M) > 0


# ---- harvester: DB layer (redirected to tmp_path) ----------------------------

@pytest.fixture
def tmp_db(tmp_path, monkeypatch):
    """Point the harvester's DB paths into tmp_path so the real DB is untouched."""
    data_dir = tmp_path / "data"
    monkeypatch.setattr(harvester, "DATA_DIR", data_dir)
    monkeypatch.setattr(harvester, "IMAGES_DIR", data_dir / "images")
    monkeypatch.setattr(harvester, "DB_PATH", data_dir / "leads.db")
    return data_dir / "leads.db"


def _sample_lead(place_id="buildings/test123"):
    return {
        "place_id": place_id,
        "address": "Testgatan 1, 268 77 Kågeröd, Sweden",
        "lat": 55.995,
        "lng": 13.085,
        "coordinates": "55.995000,13.085000",
        "solar_confidence": "HIGH",
        "roof_area_m2": 142.5,
        "image_path": "data/images/test123.png",
        "raw": {"name": place_id, "imageryQuality": "HIGH"},
    }


class TestDbLayer:
    def test_init_db_creates_db_and_dirs(self, tmp_db):
        harvester.init_db()
        assert tmp_db.exists()
        assert harvester.IMAGES_DIR.exists()
        with sqlite3.connect(tmp_db) as c:
            tables = {r[0] for r in c.execute(
                "SELECT name FROM sqlite_master WHERE type='table'")}
        assert "leads" in tables

    def test_init_db_is_idempotent(self, tmp_db):
        harvester.init_db()
        harvester.init_db()  # must not raise on existing schema
        assert tmp_db.exists()

    def test_insert_lead_and_lead_exists(self, tmp_db):
        harvester.init_db()
        assert not harvester.lead_exists("buildings/test123")
        harvester.insert_lead(_sample_lead())
        assert harvester.lead_exists("buildings/test123")
        assert not harvester.lead_exists("buildings/other")

    def test_insert_lead_persists_fields_with_pending_status(self, tmp_db):
        harvester.init_db()
        harvester.insert_lead(_sample_lead())
        with sqlite3.connect(tmp_db) as c:
            c.row_factory = sqlite3.Row
            row = c.execute("SELECT * FROM leads").fetchone()
        assert row["status"] == "pending"
        assert row["address"] == "Testgatan 1, 268 77 Kågeröd, Sweden"
        assert row["lat"] == pytest.approx(55.995)
        assert row["roof_area_m2"] == pytest.approx(142.5)
        assert row["created_at"]  # non-empty timestamp

    def test_insert_duplicate_place_id_rejected(self, tmp_db):
        harvester.init_db()
        harvester.insert_lead(_sample_lead())
        with pytest.raises(sqlite3.IntegrityError):
            harvester.insert_lead(_sample_lead())

    def test_real_db_path_not_used(self, tmp_db):
        harvester.init_db()
        assert str(harvester.DB_PATH).startswith(str(tmp_db.parent.parent))


# ---- app: pure helpers --------------------------------------------------------

class TestEstimateFuse:
    @pytest.mark.parametrize("area, expected", [
        (None, ""),
        (99, "16 A"),
        (100, "20 A"),
        (149, "20 A"),
        (150, "25 A"),
        (199, "25 A"),
        (200, "35 A"),
    ])
    def test_boundaries(self, area, expected):
        assert app.estimate_fuse(area) == expected


class TestPlusCode:
    def test_returns_non_empty_string(self):
        code = app.plus_code(55.99, 13.08)
        assert isinstance(code, str)
        assert code
        assert "+" in code  # Plus Codes always contain the '+' separator


class TestMrkollLink:
    def test_none_address_gives_empty_string(self):
        assert app.mrkoll_link(None) == ""

    def test_empty_address_gives_empty_string(self):
        assert app.mrkoll_link("") == ""

    def test_address_is_url_encoded(self):
        link = app.mrkoll_link("Testgatan 1, Kågeröd")
        assert link.startswith("https://mrkoll.se/resultat/?n=&a=")
        assert "Testgatan+1" in link
        assert " " not in link
        assert "å" not in link and "ö" not in link  # percent-encoded


class TestMapsLink:
    def test_contains_lat_lng(self):
        link = app.maps_link(55.99, 13.08)
        assert "55.99,13.08" in link
        assert link.startswith("https://www.google.com/maps/search/?api=1&query=")


# ---- app: Excel export ---------------------------------------------------------

def _fake_rows() -> list[sqlite3.Row]:
    """Real sqlite3.Row objects from an in-memory DB (matches app.py's usage)."""
    c = sqlite3.connect(":memory:")
    c.row_factory = sqlite3.Row
    c.execute("CREATE TABLE leads (address TEXT, lat REAL, lng REAL, roof_area_m2 REAL)")
    c.executemany(
        "INSERT INTO leads VALUES (?, ?, ?, ?)",
        [
            ("Testgatan 1, Kågeröd", 55.995, 13.085, 142.5),
            (None, 55.990, 13.080, None),  # missing address + area must not crash
        ],
    )
    rows = c.execute("SELECT * FROM leads").fetchall()
    c.close()
    return rows


class TestBuildExcelBytes:
    def test_returns_xlsx_bytes_with_zip_magic(self):
        data = app.build_excel_bytes(_fake_rows())
        assert isinstance(data, bytes)
        assert data[:2] == b"PK"

    def test_empty_rows_still_valid_workbook(self):
        data = app.build_excel_bytes([])
        assert data[:2] == b"PK"

    def test_content_round_trip(self, tmp_path):
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(app.build_excel_bytes(_fake_rows())))
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        assert header == app.EXPORT_COLUMNS
        first = [cell.value for cell in ws[2]]
        assert first[0] == "Testgatan 1, Kågeröd"
        assert first[2] == "20 A"  # 142.5 m2 -> 20 A
        # openpyxl reads back empty-string cells as None, so accept both.
        second = [cell.value for cell in ws[3]]
        assert second[0] in ("", None)  # None address -> ""
        assert second[2] in ("", None)  # None area -> ""
        assert second[3] in ("", None)  # None address -> no MrKoll link
