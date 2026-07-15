"""test_export_bild.py — Bild-kolumnen i Excel-exporten.

Verifierar att:
  1. _lead_image_url föredrar confirmed_image_url, faller tillbaka på
     nyckelfri LM WMS-URL från lat/lng, och aldrig ger icke-http-skräp.
  2. _to_excel_bytes gör Bild-kolumnens URL:er till riktiga klickbara
     hyperlänkar (öppningsbara direkt från Excel).
"""
import io
import sys
import types
from unittest.mock import MagicMock

import pandas as pd

# Samma streamlit-stub som test_background_scan.py
_st = types.ModuleType("streamlit")
for _attr in [
    "cache_resource", "cache_data", "session_state", "secrets", "error",
    "warning", "info", "success", "rerun", "expander", "code", "text_input",
    "selectbox", "button", "metric", "spinner", "tabs", "columns", "divider",
    "caption", "subheader", "header", "markdown", "dataframe", "image",
    "sidebar", "checkbox", "radio", "empty", "progress", "stop", "toast",
]:
    setattr(_st, _attr, MagicMock())
_st.secrets = {}
_st.session_state = MagicMock()
sys.modules.setdefault("streamlit", _st)
for _mod in ["extra_streamlit_components", "stripe", "googlemaps"]:
    sys.modules.setdefault(_mod, types.ModuleType(_mod))

import app  # noqa: E402


def test_lead_image_url_prefers_confirmed():
    row = {"confirmed_image_url": "https://ex.se/bild.jpg", "lat": 56.9, "lng": 14.8}
    assert app._lead_image_url(row) == "https://ex.se/bild.jpg"


def test_lead_image_url_falls_back_to_lm_wms():
    row = {"confirmed_image_url": None, "lat": 56.88515, "lng": 14.76547, "image_url": None}
    url = app._lead_image_url(row)
    assert url.startswith("https://minkarta.lantmateriet.se/")
    assert "GetMap" in url


def test_lead_image_url_empty_when_nothing():
    assert app._lead_image_url({"confirmed_image_url": None, "lat": None, "lng": None}) == ""


def test_excel_bild_column_is_hyperlink():
    df = pd.DataFrame({
        "Bild": ["https://minkarta.lantmateriet.se/map/ortofoto?x=1", ""],
        "Adress": ["Testgatan 1, Växjö", "Testgatan 2, Växjö"],
    })
    data = app._to_excel_bytes(df)

    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(data))["Leads"]
    cell = ws["A2"]  # första dataraden i Bild-kolumnen
    assert cell.hyperlink is not None
    assert cell.hyperlink.target.startswith("https://minkarta.lantmateriet.se/")
    assert cell.value == "Öppna bild"
    # Tom URL lämnas orörd — ingen död länk
    assert ws["A3"].hyperlink is None
