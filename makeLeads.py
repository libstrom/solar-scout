#!/usr/bin/env python3
"""
makeLeads.py — Genererar leads.xlsx optimerat för telefon-mötesbokning

Flikar:
  • Ringlista   — scorad ringköö med pitch-text, produkt-bucket, CRM-kolumner
  • Scoring     — förklaring av scoringmodellen
  • Översikt    — statistik per kontakttyp, kommun, byggår

Usage:
    python makeLeads.py enspecta.tab [leads.xlsx] [--energy energy-data.json]

Med --energy används verklig energiklass + elförbrukning från XLSM-filerna
(genereras av: node batchXlsm.mjs <mapp-med-xlsm> energy-data.json).
Utan --energy används estimat baserat på byggår.
"""
import sys, re, functools, json
from collections import Counter
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import Rule
except ImportError:
    print("Kör: pip install openpyxl")
    sys.exit(1)


# ── helpers ───────────────────────────────────────────────────────────────────

_CTRL = re.compile(r'[\x00-\x1f]')

def clean(s):
    if not s: return ''
    return _CTRL.sub(' ', s).strip()

def norm(s):
    return clean(s).lower().replace('  ', ' ')

def tc(s):
    return clean(s).title() if s else ''

def yr(s):
    m = re.match(r'(\d{4})', str(s or ''))
    return int(m.group(1)) if m else None


# ── scoring ───────────────────────────────────────────────────────────────────

def score_lead(byggår_str, kontakttyp, besdat_str, telefon, email, energy=None):
    """
    Score 0–100.
    Med energy-data (XLSM): energiklass + elförbrukning + solceller används.
    Utan energy-data: estimat baserat på byggår.
    """
    points = 0
    reasons = []

    if energy:
        # ── XLSM-baserad husålder/energi (max 45p) ──────────────────────────
        ek = (energy.get('energiklass') or '').upper()
        ek_pts = {'A': 0, 'B': 5, 'C': 10, 'D': 20, 'E': 30, 'F': 38, 'G': 45}.get(ek)
        if ek_pts is not None:
            points += ek_pts; reasons.append(f'Energiklass {ek} (+{ek_pts})')
        else:
            y = yr(byggår_str) or yr(str(energy.get('nybyggnadsar', '')))
            if y:
                p = 40 if y<=1960 else 35 if y<=1970 else 28 if y<=1980 else 18 if y<=1990 else 10 if y<=2000 else 4
                points += p; reasons.append(f'Byggt {y} (+{p})')

        # ── direktel-bonus (max +20) ─────────────────────────────────────────
        epk = energy.get('energi_per_kalla') or {}
        el_direkt = (epk.get('el_direkt') or 0) + (epk.get('el_vattenburen') or 0)
        if el_direkt > 5000:
            points += 20; reasons.append(f'Direktel {el_direkt:.0f} kWh (+20)')
        elif el_direkt > 1000:
            points += 10; reasons.append(f'El-uppvärmd (+10)')

        # ── total elförbrukning ──────────────────────────────────────────────
        total_el = energy.get('total_el_kwh') or 0
        if total_el > 25000:
            points += 8; reasons.append(f'Elförbrukning {total_el:.0f} kWh (+8)')
        elif total_el > 15000:
            points += 4; reasons.append(f'Elförbrukning {total_el:.0f} kWh (+4)')

        # ── har redan solceller (negativt) ──────────────────────────────────
        if energy.get('har_solceller'):
            points -= 20; reasons.append('Har redan solceller (-20)')

        # ── åtgärdsförslag nämner sol/batteri ───────────────────────────────
        atg = energy.get('atgardsforslag') or {}
        atg_text = (atg.get('text') or '') if isinstance(atg, dict) else str(atg)
        if re.search(r'solcell|batteri|v[äa]rmepump', atg_text, re.I):
            points += 8; reasons.append('Åtgärdsförslag: sol/batteri/VP (+8)')

    else:
        # ── Estimat baserat på byggår (ingen XLSM) ───────────────────────────
        y = yr(byggår_str)
        if y:
            if y <= 1960:   points += 40; reasons.append(f'Byggt {y} (+40)')
            elif y <= 1970: points += 35; reasons.append(f'Byggt {y} (+35)')
            elif y <= 1980: points += 28; reasons.append(f'Byggt {y} (+28)')
            elif y <= 1990: points += 18; reasons.append(f'Byggt {y} (+18)')
            elif y <= 2000: points += 10; reasons.append(f'Byggt {y} (+10)')
            else:           points += 4;  reasons.append(f'Byggt {y} (+4)')
        else:
            points += 15; reasons.append('Byggår okänt (+15, antas äldre)')

    # ── Kontakttyp (samma oavsett XLSM) ─────────────────────────────────────
    if kontakttyp == 'Köpare':
        points += 30; reasons.append('Köpare/nuv. ägare (+30)')
    elif kontakttyp == 'Intressent':
        points += 20; reasons.append('Intressent/visning (+20)')
    else:
        points += 8;  reasons.append('Säljare/har flyttat (+8)')

    # ── Besiktningsdatum ─────────────────────────────────────────────────────
    bd = yr(besdat_str)
    if bd:
        age = date.today().year - bd
        if age <= 2:   points += 20; reasons.append(f'Besiktning {bd} (+20)')
        elif age <= 5: points += 12; reasons.append(f'Besiktning {bd} (+12)')
        elif age <= 10:points += 6;  reasons.append(f'Besiktning {bd} (+6)')

    # ── Kontaktinfo ──────────────────────────────────────────────────────────
    if telefon: points += 8; reasons.append('Har telefon (+8)')
    if email:   points += 4; reasons.append('Har e-post (+4)')

    return min(max(points, 0), 100), ' | '.join(reasons)


def bucket(score, byggår_str, energy=None):
    har_sol = energy.get('har_solceller') if energy else False
    if har_sol:
        return '🔋 BATTERI (har sol)'
    y = yr(byggår_str)
    if score >= 65:
        return '☀️🔋 SOL + BATTERI'
    if score >= 45:
        if y and y <= 2005:
            return '☀️🔋 SOL + BATTERI'
        return '☀️ SOL'
    return '🔋 BATTERI / VP'


def score_color(score):
    if score >= 70: return 'B71C1C'   # mörkröd = het
    if score >= 55: return 'E65100'   # orange
    if score >= 40: return 'F9A825'   # gul
    return '555555'                    # grå


def pitch_text(namn, adress, ort, byggår_str, bucket_str, energy=None):
    """Färdig pitch-text för David att klistra in i mötesbok."""
    y = yr(byggår_str)
    fornamn = namn.split()[0] if namn else 'du'
    har_sol = energy.get('har_solceller') if energy else False
    ek      = (energy.get('energiklass') or '').upper() if energy else ''
    epk     = (energy.get('energi_per_kalla') or {}) if energy else {}
    el_dir  = (epk.get('el_direkt') or 0) + (epk.get('el_vattenburen') or 0)

    if har_sol:
        behov   = "du har redan solceller — nu är det rätt tid att lägga till batteri"
        produkt = "batterilagring som gör att du kan spara och använda din egenproducerade el på kvällen"
    elif ek in ('F', 'G'):
        behov   = f"fastigheten har energiklass {ek} — det finns stor besparingspotential"
        produkt = "sol + batteri + möjlig VP — vi kan halvera er energikostnad"
    elif el_dir > 5000:
        behov   = "huset värms med el — det är det dyraste uppvärmningssättet just nu"
        produkt = "sol + batteri — du kan kapa elräkningen med upp till 70%"
    elif y and y <= 1980:
        behov   = f"hus från {y}-talet har ofta höga elkostnader och stor potential"
        produkt = "sol + batteri — du kan kapa elräkningen med upp till 70%"
    elif y and y <= 1995:
        behov   = f"villaägare i {ort or 'ert område'} väljer allt oftare solceller"
        produkt = "solceller med batteri — du lagrar överskottet och säljer resten"
    else:
        behov   = "många väljer nu att komplettera med batteri för att maximera egenkonsumtion"
        produkt = "batterilagring — perfekt om du redan har sol eller funderar på det"

    return (
        f"Hej {fornamn}! Jag heter [säljare] och ringer från Enspecta Ensolar. "
        f"Vi hjälper villaägare i {ort or 'Sverige'} att sänka sina energikostnader — "
        f"{behov}. "
        f"Vi erbjuder {produkt}. "
        f"Får jag boka in ett kostnadsfritt hembesök på {adress or 'er fastighet'} "
        f"så går vi igenom er potential tillsammans?"
    )


# ── parse enspecta.tab ────────────────────────────────────────────────────────

def parse_tab(path):
    with open(path, encoding='utf-8', errors='replace') as f:
        lines = [l.rstrip('\r\n') for l in f]

    by_fastig = {}
    cur_fastig = cur_status = cur_rec = None
    cur_ints = []

    def flush():
        if not cur_fastig or not cur_rec: return
        if cur_fastig not in by_fastig:
            by_fastig[cur_fastig] = {'kopare': [], 'intressenter': [], 'saljare': []}
        b = by_fastig[cur_fastig]
        if cur_status == 'Köpare':    b['kopare'].append(cur_rec)
        elif cur_status == 'Säljare': b['saljare'].append(cur_rec)
        if cur_status == 'Säljare' and cur_ints:
            b['intressenter'].extend(cur_ints)

    def mk_int(c, prop):
        fn = clean(c[24]) if len(c)>24 else ''
        if not fn: return None
        return {
            'namn':    f"{fn} {clean(c[25]) if len(c)>25 else ''}".strip(),
            'telefon': (clean(c[27]) if len(c)>27 else '') or (clean(c[28]) if len(c)>28 else ''),
            'email':   clean(c[26]) if len(c)>26 else '',
            'adress':  prop.get('adress','') if prop else '',
            'postnr':  prop.get('postnr','') if prop else '',
            'ort':     prop.get('ort','')    if prop else '',
            'kommun':  prop.get('kommun','') if prop else '',
            'besdat':  prop.get('besdat','') if prop else '',
            'byggår':  prop.get('byggår','') if prop else '',
        }

    for line in lines:
        c = line.split('\t')
        case_id = clean(c[0]) if c else ''
        if case_id:
            flush()
            fastig = norm(c[12]) if len(c)>12 else ''
            cur_fastig = fastig or None
            cur_status = clean(c[1]) if len(c)>1 else ''
            cur_rec = {
                'case_id': case_id,
                'namn':    clean(c[9])  if len(c)>9  else '',
                'namn2':   clean(c[20]) if len(c)>20 else '',
                'adress':  clean(c[5])  if len(c)>5  else '',
                'postnr':  clean(c[6])  if len(c)>6  else '',
                'ort':     clean(c[7])  if len(c)>7  else '',
                'kommun':  clean(c[8])  if len(c)>8  else '',
                'telefon': clean(c[23]) if len(c)>23 else '',
                'email':   clean(c[22]) if len(c)>22 else '',
                'besdat':  clean(c[2])  if len(c)>2  else '',
                'byggår':  clean(c[18]).replace('--','') if len(c)>18 else '',
                'renovat': clean(c[19]).replace('--','') if len(c)>19 else '',
                'fastig':  fastig,
            } if fastig else None
            cur_ints = []
            i = mk_int(c, cur_rec)
            if i: cur_ints.append(i)
        else:
            fn = clean(c[24]) if len(c)>24 else ''
            if fn and cur_rec:
                i = mk_int(c, cur_rec)
                if i: cur_ints.append(i)
    flush()
    return by_fastig


def best_contact(b):
    for r in b['kopare']:
        if r['telefon'] or r['email']:
            return 'Köpare', r, b['intressenter']
    for r in b['intressenter']:
        if r.get('telefon') or r.get('email'):
            prop = b['saljare'][0] if b['saljare'] else {}
            merged = {**prop, **r, 'namn': r['namn'], 'namn2': '',
                      'case_id': prop.get('case_id',''),
                      'besdat':  r.get('besdat', prop.get('besdat','')),
                      'byggår':  r.get('byggår',  prop.get('byggår',''))}
            return 'Intressent', merged, []
    for r in b['saljare']:
        if r['telefon'] or r['email']:
            return 'Säljare', r, b['intressenter']
    return None, None, []


# ── Excel styling ─────────────────────────────────────────────────────────────

@functools.lru_cache(maxsize=64)
def F(h):  return PatternFill('solid', fgColor=h)
def Fn(*a,**k): return Font(*a,**k)

@functools.lru_cache(maxsize=8)
def border(color='DDDDDD'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

BLU='1A237E'; BLU2='283593'; BLU3='E8EAF6'
GRN='E8F5E9'; GRN2='1B5E20'
YLW='FFFDE7'; YLW2='F57F17'
ORG='FFF3E0'; ORG2='E65100'
RED='FFEBEE'; RED2='B71C1C'
HOT='FFF8E1'  # score ≥ 70

def score_bar(score):
    filled = round(score / 10)
    return '█' * filled + '░' * (10 - filled)

STATUS_OPTS  = '"Ej kontaktad,Bokad möte,Ej intresserad,Återkom,Fel nummer,Röstbrevlåda,Såld"'
SALJARE_OPTS = '"Ivan,David,Linus,Annan"'

# (header, width, col-index-0-based)
COLS = [
    ('Score',              7),   # 0
    ('Bucket',            18),   # 1
    ('Status',            16),   # 2
    ('Säljare',           10),   # 3
    ('Nästa kontakt',     14),   # 4
    ('Namn',              26),   # 5
    ('Telefon',           16),   # 6
    ('E-post',            30),   # 7
    ('Adress',            28),   # 8
    ('Postnr',             8),   # 9
    ('Ort',               16),   # 10
    ('Kommun',            16),   # 11
    ('Byggår',             8),   # 12
    ('Energiklass',        9),   # 13  ← ny
    ('Kontakttyp',        13),   # 14
    ('Namn 2',            20),   # 15
    ('Intressent — Namn', 22),   # 16
    ('Intressent — Tel',  16),   # 17
    ('Pitch-text',        55),   # 18
    ('Score-förklaring',  45),   # 19
    ('Besiktningsdatum',  16),   # 20
    ('Renoverat',          9),   # 21
    ('Fastighetsbeteckning', 24),# 22
    ('Anteckningar',      35),   # 23
    ('CaseID',            12),   # 24
]


TOP_COLS = [
    ('Score',       7),
    ('Bar',        12),
    ('Bucket',     20),
    ('Namn',       26),
    ('Telefon',    16),
    ('Adress',     28),
    ('Ort',        16),
    ('Energiklass', 9),
    ('Kontakttyp', 13),
    ('Pitch-text', 60),
    ('Anteckningar', 35),
]
# indices into main rows tuple matching TOP_COLS
TOP_IDX = [0, None, 1, 5, 6, 8, 10, 13, 14, 18, 23]


def make_top50(wb, rows, n=50):
    ws = wb.create_sheet('🔥 TOP 50', 0)
    ws.sheet_view.showGridLines = False
    top = rows[:n]
    LC = get_column_letter(len(TOP_COLS))

    # Banner
    ws.merge_cells(f'A1:{LC}1')
    b = ws['A1']
    b.value = f'🔥  ENSPECTA TOP {n}  —  De hetaste leadsen just nu  •  {date.today()}'
    b.font  = Fn(bold=True, size=13, color='FFFFFF')
    b.fill  = F('B71C1C')
    b.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    # Sub-banner
    ws.merge_cells(f'A2:{LC}2')
    g = ws['A2']
    g.value = 'Ring uppifrån och ned  •  Status dropdowns i kolumn A (välj flik Ringlista för komplett lista)'
    g.font  = Fn(size=10, italic=True, color='B71C1C')
    g.fill  = F('FFEBEE')
    g.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18

    # Headers
    for ci, (h, _) in enumerate(TOP_COLS, 1):
        c = ws.cell(3, ci, h)
        c.fill = F('B71C1C')
        c.font = Fn(bold=True, color='FFFFFF', size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border('C62828')
    ws.row_dimensions[3].height = 22

    _bkt_fills = {'☀️🔋 SOL + BATTERI': (F('FFF8E1'), Fn(bold=True, color='E65100', size=10)),
                  '☀️ SOL':              (F('E8F5E9'), Fn(bold=True, color='1B5E20', size=10)),
                  '🔋 BATTERI / VP':     (F('E3F2FD'), Fn(bold=True, color='1565C0', size=10)),
                  '🔋 BATTERI (har sol)':(F('E3F2FD'), Fn(bold=True, color='1565C0', size=10))}
    ek_colors = {'A':'1B5E20','B':'2E7D32','C':'388E3C','D':'F9A825','E':'E65100','F':'C62828','G':'B71C1C'}
    _score_fonts = {c: Fn(bold=True, size=13, color=c) for c in ('B71C1C','E65100','F9A825','555555')}
    _border_std = border()

    for ri, row in enumerate(top, 4):
        sc  = row[0]
        hot = sc >= 70
        row_fill = F('FFF8E1') if sc >= 70 else F('FFF3E0') if sc >= 55 else F('FAFAFA')

        for ci, (col_info, src_idx) in enumerate(zip(TOP_COLS, TOP_IDX), 1):
            val = score_bar(sc) if src_idx is None else (row[src_idx] if src_idx < len(row) else '')
            c = ws.cell(ri, ci, val)
            c.fill   = row_fill
            c.border = _border_std
            c.alignment = Alignment(vertical='center', wrap_text=False)

            if ci == 1:  # Score
                c.font = _score_fonts[score_color(sc)]
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 2:  # Bar
                c.font = Fn(name='Consolas', size=10, color=score_color(sc))
                c.alignment = Alignment(horizontal='left', vertical='center')
            elif ci == 3:  # Bucket
                bfill, bfont = _bkt_fills.get(val, (row_fill, Fn(bold=True, size=10)))
                c.fill = bfill; c.font = bfont
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 8:  # Energiklass
                c.font = Fn(bold=True, size=12, color=ek_colors.get(val or '', '555555'))
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 10:  # Pitch-text
                c.font = Fn(size=9, color='333333')
                c.alignment = Alignment(vertical='center', wrap_text=True)
            elif ci == 11:  # Anteckningar
                c.value = ''

        ws.row_dimensions[ri].height = 38 if hot else 24

    tbl = Table(displayName='Top50', ref=f'A3:{LC}{len(top)+3}')
    tbl.tableStyleInfo = TableStyleInfo(name='TableStyleLight2',
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)
    ws.freeze_panes = 'A4'
    for ci, (_, w) in enumerate(TOP_COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def make_ringlista(wb, rows):
    ws = wb.active
    ws.title = 'Ringlista'
    ws.sheet_view.showGridLines = False
    ncols = len(COLS)
    LC = get_column_letter(ncols)

    # Banner
    ws.merge_cells(f'A1:{LC}1')
    b = ws['A1']
    b.value = ('☀️  ENSPECTA LEADS — Telefon-mötesbokning  |  '
               'Sortera Score högt→lågt  •  Filtrera Ort för daglig körning  •  '
               'Pitch-text i kolumn R — klistra in i mötesbok')
    b.font = Fn(bold=True, size=11, color='FFFFFF')
    b.fill = F(BLU)
    b.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Guide-rad
    ws.merge_cells(f'A2:{LC}2')
    g = ws['A2']
    g.value = ('FILTER:  Score ▾ → sortera fallande  •  Status = "Ej kontaktad"  •  '
               'Bucket = "☀️🔋 SOL + BATTERI"  •  Kommun = valfritt område')
    g.font  = Fn(size=10, color=BLU, italic=True)
    g.fill  = F(BLU3)
    g.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 20

    # Headers
    for ci, (h, _) in enumerate(COLS, 1):
        c = ws.cell(3, ci, h)
        c.fill = F(BLU2)
        c.font = Fn(bold=True, color='FFFFFF', size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border('1A237E')
    ws.row_dimensions[3].height = 24

    # Validations
    dv_status  = DataValidation(type='list', formula1=STATUS_OPTS,  showDropDown=False, showErrorMessage=False)
    dv_saljare = DataValidation(type='list', formula1=SALJARE_OPTS, showDropDown=False, showErrorMessage=False)
    dv_status.sqref  = f'C4:C{len(rows)+3}'
    dv_saljare.sqref = f'D4:D{len(rows)+3}'
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_saljare)

    TYPE_FILL = {'Köpare': F(GRN), 'Intressent': F(YLW), 'Säljare': F(ORG)}
    TYPE_FILL_HOT = {'Köpare': F('C8E6C9'), 'Intressent': F('FFF176'), 'Säljare': F('FFCC80')}

    # Pre-create shared style objects — avoids re-allocating per cell
    _border_std  = border()
    _align_std   = Alignment(vertical='center', wrap_text=False)
    _align_ctr   = Alignment(horizontal='center', vertical='center')
    _align_wrap  = Alignment(vertical='center', wrap_text=True)
    _font_status = Fn(color='777777', italic=True, size=10)
    _font_mono   = Fn(name='Consolas', size=10)
    _font_pitch  = Fn(size=9, color='333333')
    _font_why    = Fn(size=9, color='777777', italic=True)
    _bkt_fills   = {'☀️🔋 SOL + BATTERI': (F('FFF8E1'), Fn(bold=True, color='E65100', size=10)),
                    '☀️ SOL':              (F('E8F5E9'), Fn(bold=True, color='1B5E20', size=10)),
                    '🔋 BATTERI / VP':     (F('E3F2FD'), Fn(bold=True, color='1565C0', size=10))}
    _typ_fonts   = {'Köpare':     Fn(bold=True, size=9, color=GRN2),
                    'Intressent': Fn(bold=True, size=9, color=YLW2),
                    'Säljare':    Fn(bold=True, size=9, color=ORG2)}
    _score_fonts = {c: Fn(bold=True, size=12, color=c)
                    for c in ('B71C1C','E65100','F9A825','555555')}

    for ri, row in enumerate(rows, 4):
        sc   = row[0]
        typ  = row[14]   # Kontakttyp at index 14
        hot  = sc >= 70
        rfill = TYPE_FILL_HOT.get(typ, F('FFFDE7')) if hot else TYPE_FILL.get(typ, F('FAFAFA'))
        sc_font = _score_fonts[score_color(sc)]
        typ_font = _typ_fonts.get(typ, Fn(bold=True, size=9))

        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.fill      = rfill
            c.border    = _border_std
            c.alignment = _align_std

            if ci == 1:    # Score — show number + bar
                c.value = f'{sc}\n{score_bar(sc)}'
                c.font = sc_font
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            elif ci == 2:  # Bucket
                bfill, bfont = _bkt_fills.get(val, (F('FFFFFF'), Fn(bold=True, size=10)))
                c.fill = bfill
                c.font = bfont
                c.alignment = _align_ctr
            elif ci == 3:  # Status default
                c.value = 'Ej kontaktad'
                c.font  = _font_status
            elif ci in (7, 8):  # Tel/email
                c.font = _font_mono
            elif ci == 14:  # Energiklass
                ek_colors = {'A':'1B5E20','B':'2E7D32','C':'388E3C','D':'F9A825','E':'E65100','F':'C62828','G':'B71C1C'}
                c.font = Fn(bold=True, size=11, color=ek_colors.get(val or '', '555555'))
                c.alignment = _align_ctr
            elif ci == 15:  # Kontakttyp
                c.font = typ_font
            elif ci == 19:  # Pitch-text
                c.font = _font_pitch
                c.alignment = _align_wrap
            elif ci == 20:  # Score-förklaring
                c.font = _font_why

        ws.row_dimensions[ri].height = 32 if hot else 20

    # Excel Table
    tbl = Table(displayName='Leads', ref=f'A3:{LC}{len(rows)+3}')
    tbl.tableStyleInfo = TableStyleInfo(
        name='TableStyleLight1',
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=False,  showColumnStripes=False,
    )
    ws.add_table(tbl)

    # Conditional: Bokad möte → grön
    ds = DifferentialStyle(fill=F('C8E6C9'))
    ws.conditional_formatting.add(f'A4:{LC}{len(rows)+3}',
        Rule(type='expression', dxf=ds, formula=['$C4="Bokad möte"']))

    # Conditional: Ej intresserad → grå
    ds2 = DifferentialStyle(fill=F('EEEEEE'), font=Fn(color='AAAAAA'))
    ws.conditional_formatting.add(f'A4:{LC}{len(rows)+3}',
        Rule(type='expression', dxf=ds2, formula=['$C4="Ej intresserad"']))

    ws.freeze_panes = 'A4'
    for ci, (_, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def make_scoring(wb):
    ws = wb.create_sheet('Scoring-modell')
    ws.sheet_view.showGridLines = False

    def hdr(r, txt, size=13):
        c = ws.cell(r, 1, txt)
        c.font = Fn(bold=True, size=size, color=BLU)

    def row3(r, factor, pts, note):
        ws.cell(r, 1, factor).font = Fn(bold=True, size=10)
        ws.cell(r, 2, pts).font    = Fn(size=10, color=RED2)
        ws.cell(r, 2).alignment    = Alignment(horizontal='right')
        ws.cell(r, 3, note).font   = Fn(size=10, color='444444', italic=True)

    r = 1
    hdr(r, '☀️  Scoring-modell — så beräknas poängen', 15); r += 2
    ws.cell(r, 1, 'Obs: Score är ett estimat baserat på enspecta.tab. '
                   'Med XLSM-data (energiklass, uppvärmning, elförbrukning) '
                   'kan precision förbättras till ±5p.').font = Fn(size=10, italic=True, color='777777')
    r += 2

    hdr(r, 'Husålder (max 40p)'); r += 1
    for factor, pts, note in [
        ('Byggd ≤ 1960', '+40', 'Troligen direktel/olja — störst behov, bäst pitch'),
        ('Byggd 1961–1970', '+35', 'Ofta Energiklass E–G'),
        ('Byggd 1971–1980', '+28', 'Energiklass D–F sannolikt'),
        ('Byggd 1981–1990', '+18', 'Blandad potential'),
        ('Byggd 1991–2000', '+10', 'Lägre behov, men sol kan löna sig'),
        ('Byggd > 2000',    '+4',  'Nyare hus — fråga om sol redan finns'),
        ('Byggår saknas',   '+15', 'Konservativt antagande — troligen äldre'),
    ]:
        row3(r, factor, pts, note); r += 1
    r += 1

    hdr(r, 'Kontakttyp (max 30p)'); r += 1
    for f, p, n in [
        ('Köpare',     '+30', 'Bor på fastigheten nu — direkt beslutare'),
        ('Intressent', '+20', 'Var på visning — troligen ny ägare'),
        ('Säljare',    '+8',  'Har flyttat — lägre sannolikhet'),
    ]:
        row3(r, f, p, n); r += 1
    r += 1

    hdr(r, 'Besiktningsdatum (max 20p)'); r += 1
    for f, p, n in [
        ('≤ 2 år sedan', '+20', 'Nytt ägarbyte — varmt lead'),
        ('3–5 år sedan', '+12', 'Relativt nytt'),
        ('6–10 år sedan','+6',  'Äldre besiktning'),
    ]:
        row3(r, f, p, n); r += 1
    r += 1

    hdr(r, 'Kontaktinfo'); r += 1
    for f, p, n in [
        ('Har telefonnummer', '+8', 'Kan ringas direkt'),
        ('Har e-postadress',  '+4', 'Kan mailas som komplement'),
    ]:
        row3(r, f, p, n); r += 1
    r += 2

    hdr(r, 'Produktbuckets'); r += 1
    for f, p, n in [
        ('☀️🔋 SOL + BATTERI', 'Score ≥ 65', 'Prioritera dessa — störst affär'),
        ('☀️ SOL',             'Score 45–64', 'Bra kandidat, fokus på sol'),
        ('🔋 BATTERI / VP',    'Score < 45',  'Fokus på batteri eller VP-tillägg'),
    ]:
        ws.cell(r, 1, f).font   = Fn(bold=True, size=10)
        ws.cell(r, 2, p).font   = Fn(size=10, color=RED2)
        ws.cell(r, 3, n).font   = Fn(size=10, italic=True, color='444444')
        r += 1

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 52


def make_oversikt(wb, rows):
    ws = wb.create_sheet('Översikt')
    ws.sheet_view.showGridLines = False

    def hdr(r, txt, size=13):
        ws.cell(r, 1, txt).font = Fn(bold=True, size=size, color=BLU)

    def kv(r, k, v, vc=None):
        ws.cell(r, 1, k).font = Fn(bold=True, size=11)
        c = ws.cell(r, 2, v)
        c.font = Fn(bold=True, size=11, color=vc or '000000')
        c.alignment = Alignment(horizontal='right')

    r = 1
    hdr(r, f'☀️  Enspecta Leads — Översikt  ({date.today()})', 15); r += 2

    stats = Counter(row[14] for row in rows)
    hdr(r, 'Kontakttyp'); r += 1
    kv(r, '🟢  Köpare',      stats['Köpare'],     GRN2); r += 1
    kv(r, '🟡  Intressent',  stats['Intressent'], YLW2); r += 1
    kv(r, '🟠  Säljare',     stats['Säljare'],    ORG2); r += 1
    kv(r, '    Totalt',       sum(stats.values()), BLU);  r += 2

    score_dist = Counter()
    for row in rows:
        s = row[0]
        if s >= 70:   score_dist['≥ 70 (het)'] += 1
        elif s >= 55: score_dist['55–69 (varm)'] += 1
        elif s >= 40: score_dist['40–54 (ok)'] += 1
        else:         score_dist['< 40 (kall)'] += 1
    hdr(r, 'Score-distribution'); r += 1
    total_r = sum(score_dist.values()) or 1
    score_colors = {'≥ 70 (het)': 'B71C1C', '55–69 (varm)': 'E65100', '40–54 (ok)': 'F9A825', '< 40 (kall)': '888888'}
    for label in ['≥ 70 (het)', '55–69 (varm)', '40–54 (ok)', '< 40 (kall)']:
        cnt = score_dist[label]
        pct = cnt / total_r
        bar = '█' * round(pct * 30)
        kv(r, label, cnt, score_colors[label])
        bc = ws.cell(r, 3, f'{bar}  {pct:.0%}')
        bc.font = Fn(name='Consolas', size=10, color=score_colors[label])
        r += 1
    ws.column_dimensions['C'].width = 40
    r += 1

    bucket_dist = Counter(row[1] for row in rows)  # col 1 = Bucket
    hdr(r, 'Produktbuckets'); r += 1
    for bk, cnt in bucket_dist.most_common():
        kv(r, bk, cnt); r += 1
    r += 1

    hdr(r, 'Topp 20 kommuner'); r += 1
    ws.cell(r, 1, 'Kommun').font = Fn(bold=True)
    ws.cell(r, 2, 'Antal').font  = Fn(bold=True)
    ws.cell(r, 2).alignment = Alignment(horizontal='right')
    r += 1
    for kom, cnt in Counter(row[11] for row in rows if row[11]).most_common(20):  # col 11 = Kommun
        ws.cell(r, 1, kom)
        ws.cell(r, 2, cnt).alignment = Alignment(horizontal='right')
        r += 1
    r += 1

    hdr(r, 'Byggår per decennium'); r += 1
    ws.cell(r, 1, 'Decennium').font = Fn(bold=True)
    ws.cell(r, 2, 'Antal').font     = Fn(bold=True)
    r += 1
    dcnt = Counter()
    for row in rows:
        y = yr(row[12])   # col 12 = Byggår
        dcnt[(y//10)*10 if y else 'Okänt'] += 1
    for dec in sorted(d for d in dcnt if d != 'Okänt'):
        ws.cell(r, 1, f'{dec}-talet')
        ws.cell(r, 2, dcnt[dec]).alignment = Alignment(horizontal='right')
        r += 1
    if 'Okänt' in dcnt:
        ws.cell(r, 1, 'Okänt')
        ws.cell(r, 2, dcnt['Okänt']).alignment = Alignment(horizontal='right')

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 12


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]
    # Parse flags
    energy_path = None
    positional  = []
    i = 0
    while i < len(args):
        if args[i] == '--energy' and i + 1 < len(args):
            energy_path = args[i + 1]; i += 2
        else:
            positional.append(args[i]); i += 1

    if not positional:
        print('Usage: python makeLeads.py enspecta.tab [leads.xlsx] [--energy energy-data.json]')
        sys.exit(1)

    tab_path = positional[0]
    out_path = positional[1] if len(positional) > 1 else tab_path.replace('.tab', '-leads.xlsx')

    # Load energy index (fastighetsbeteckning → energy record)
    energy_index = {}
    if energy_path:
        with open(energy_path, encoding='utf-8') as f:
            energy_index = json.load(f)
        print(f'Energidata: {len(energy_index)} fastigheter från {energy_path}')
    else:
        print('Obs: kör med --energy energy-data.json för exakt scoring (node batchXlsm.mjs ...)')

    print(f'Läser {tab_path} ...')
    by_fastig = parse_tab(tab_path)
    print(f'  {len(by_fastig)} unika fastigheter')

    print('Scorar och bygger rader ...')
    rows = []
    stats = Counter()
    energy_matched = 0

    for fastig, b in by_fastig.items():
        typ, r, ints = best_contact(b)
        if not r: continue
        stats[typ] += 1
        int0 = ints[0] if ints and typ != 'Intressent' else {}

        bår     = r.get('byggår', '')
        besdat  = r.get('besdat','') or r.get('besiktningsdatum','')
        tel     = r.get('telefon','')
        epost   = r.get('email','')
        ort_str = tc(r.get('ort',''))
        adr_str = r.get('adress','')
        namn    = r.get('namn','')

        energy  = energy_index.get(fastig)   # fastig is already normalized
        if energy: energy_matched += 1

        sc, sc_why = score_lead(bår, typ, besdat, tel, epost, energy)
        bkt        = bucket(sc, bår, energy)
        pitch      = pitch_text(namn, adr_str, ort_str, bår, bkt, energy)

        ek_str = (energy.get('energiklass') or '') if energy else ''
        rows.append([
            sc,                                                          # 0  Score
            bkt,                                                         # 1  Bucket
            '',                                                          # 2  Status
            '',                                                          # 3  Säljare
            '',                                                          # 4  Nästa kontakt
            namn,                                                        # 5  Namn
            tel,                                                         # 6  Telefon
            epost,                                                       # 7  E-post
            adr_str,                                                     # 8  Adress
            r.get('postnr',''),                                          # 9  Postnr
            ort_str,                                                     # 10 Ort
            tc(r.get('kommun','')),                                      # 11 Kommun
            bår,                                                         # 12 Byggår
            ek_str,                                                      # 13 Energiklass
            typ,                                                         # 14 Kontakttyp
            r.get('namn2','') if r.get('namn2')!=namn else '',           # 15 Namn 2
            int0.get('namn',''),                                         # 16 Intressent namn
            int0.get('telefon',''),                                      # 17 Intressent tel
            pitch,                                                       # 18 Pitch-text
            sc_why,                                                      # 19 Score-förklaring
            besdat,                                                      # 20 Besiktningsdatum
            r.get('renovat',''),                                         # 21 Renoverat
            tc(fastig),                                                  # 22 Fastighetsbeteckning
            '',                                                          # 23 Anteckningar
            r.get('case_id',''),                                         # 24 CaseID
        ])

    rows.sort(key=lambda r: -r[0])   # Score fallande

    print(f'Bygger Excel ({len(rows)} rader, 4 flikar) ...')
    wb = Workbook()
    make_ringlista(wb, rows)   # uses wb.active (default sheet) — must run before any insert
    make_top50(wb, rows)       # inserts at index 0 → becomes first tab
    make_scoring(wb)
    make_oversikt(wb, rows)
    wb.save(out_path)

    hot   = sum(1 for r in rows if r[0] >= 70)
    warm  = sum(1 for r in rows if 55 <= r[0] < 70)
    sb    = sum(1 for r in rows if r[1] == '☀️🔋 SOL + BATTERI')

    print(f'\n=== Klar ===')
    print(f'  Köpare:           {stats["Köpare"]}')
    print(f'  Intressent:       {stats["Intressent"]}')
    print(f'  Säljare:          {stats["Säljare"]}')
    if energy_index:
        print(f'  Med energidata:   {energy_matched} / {len(rows)} ({energy_matched*100//max(len(rows),1)}% träff)')
    print(f'  Score ≥ 70 (het): {hot}')
    print(f'  Score 55–69:      {warm}')
    print(f'  SOL+BATTERI:      {sb}')
    print(f'\nSparad till: {out_path}')
    print('Tips: Sortera Score fallande → de hetaste leadsen överst.')


if __name__ == '__main__':
    main()
